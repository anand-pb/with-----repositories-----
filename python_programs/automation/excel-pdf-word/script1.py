from openpyxl import load_workbook

wb = load_workbook('read-example.xlsx')

print(wb.sheetnames)

ws = wb.active

for row in ws.values:
    for value in row:
        print(value)

