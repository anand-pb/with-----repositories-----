from openpyxl import Workbook

import os

os.chdir('F:\\python_programs\\automation\\excel-word-pdf')

wb = Workbook()

wb.save('write-example.xlsx')

ws = wb.active

ws = wb.create_sheet("sheet-one")

wb.save('write-example.xlsx')

ws['A4'] = 34

wb.save('write-example.xlsx')

